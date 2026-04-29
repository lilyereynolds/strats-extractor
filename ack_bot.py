"""
PPM Ack Letter Support Bot
Upload: (1) reference deal Excel, (2) new deal Excel, (3) new deal PDF.
Outputs the new deal Excel with PPM Language (E), PPM Number (F), Etie formula (G), Precision (H) filled.
"""

import io
import re
import pdfplumber
import openpyxl
import streamlit as st

# ── PDF helpers ───────────────────────────────────────────────────────────────

def extract_pdf_text(pdf_bytes: bytes) -> str:
    parts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                parts.append(t)
    return " ".join(parts)


def find_passage(text: str, keywords: list, window: int = 700, start_at_kw: bool = False) -> str | None:
    tl = text.lower()
    for kw in keywords:
        i = tl.find(kw.lower())
        if i != -1:
            if start_at_kw:
                start = i
            else:
                # Walk back to nearest sentence boundary
                look_back = text[max(0, i - 300): i]
                m = list(re.finditer(r'[.!?]\s+|\n\s*\n', look_back))
                if m:
                    start = max(0, i - 300) + m[-1].end()
                else:
                    start = max(0, i - 100)
            end = min(len(text), i + window)
            return text[start:end].strip()
    return None


def find_pct_passage(text: str, col_c_pct: float, context_kws: list, window: int = 700):
    """
    Find 'approximately X.XX%' matching col_c_pct anywhere in the PDF,
    verify one of context_kws appears within 600 chars of the match,
    return (passage, decimal_value, precision).
    Two-column PDFs break long keyword strings, so we search by value first.
    """
    for m in re.finditer(r'approximately\s+([\d]+(?:\.[\d]+)?)\s*%', text, re.I):
        s = m.group(1)
        v = float(s)
        if abs(v - col_c_pct) < 0.5:
            neighbourhood = text[max(0, m.start() - 600): m.end() + 200].lower()
            if any(kw.lower() in neighbourhood for kw in context_kws):
                prec = len(s.split('.')[1]) if '.' in s else 0
                p_start = max(0, m.start() - 300)
                p_end = min(len(text), m.end() + 400)
                return text[p_start:p_end].strip(), v / 100.0, prec
    return None, None, None


# ── Reference map ─────────────────────────────────────────────────────────────

def build_ref_map(wb: openpyxl.Workbook) -> dict:
    ws = wb.active
    m = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        desc = row[1]
        ppm_lang = row[4]
        etie = row[6]
        if desc and ppm_lang and desc not in m:
            m[desc] = {"lang": str(ppm_lang), "etie": etie}
    return m


# ── Row classification ────────────────────────────────────────────────────────

# Rows where we extract a specific number from PDF and build a formula
FORMULA_ROWS = {
    "% Mortgage Assets with CTC",
    "% Mortgage Assets with US",
    "% Mortgage Assets with DB",
}

# Keyword lists to find the right passage in the PDF for each description
SEARCH_KW = {
    # context keywords used by find_pct_passage (shorter = more robust against 2-column PDFs)
    "% Mortgage Assets with CTC":          ["Computershare", "CTC Custodian"],
    "% Mortgage Assets with US":           ["U.S. Bank", "U.S. Bank Custodian"],
    "% Mortgage Assets with DB":           ["Deutsche Bank"],
    # passage keywords used by find_passage
    "REO Loans":                           ["Initial REO Properties"],
    "% Servicer - Selene":                 ["will service all of the Mortgage Assets", "Selene Base Servicing Agreement"],
    "% Servicer - Fay":                    ["Fay Servicing", "Fay will service"],
    "Number of Mortgage Loans":            ["Number of Mortgage Loans"],
    "Aggregate UPB of Mortgage Loans":     ["Aggregate Unpaid Principal Balance"],
    "Number of REO Properties":            ["Number of REO Properties"],
    "Aggregate UPB of REO Properties":     ["Aggregate Unpaid Principal Balance of the REO"],
}

# Pool description passage covers all these rows
POOL_DESC_ROWS = {
    "Step Rate loans (REO excluded)",
    "ARM Loans (REO excluded)",
    "Fixed Loans (REO excluded)",
    "Balloon % (REO excluded)",
    "IO % (REO excluded)",
    "Amortizing % (REO excluded)",
    "% Performing",
    "Mortgage Loans Status (REO excluded)",
    "Lien Type (REO excluded)",
    "Property Type (REO excluded)",
    "Min Loan Age (REO excluded)",
}

POOL_KW = ["a pool of seasoned", "fixed-rate, step-rate"]

# Rows where the PPM language is a fixed standard phrase (not worth searching the PDF)
FIXED_LANGUAGE = {
    "Stated Final Maturity Date": "Stated Final Maturity Date",
}

# Rows to skip entirely
SKIP_ROWS = {"Summary Information", "Sanity Check", "PPM Language", "PPM Number", "Etie"}


# ── Main processor ────────────────────────────────────────────────────────────

def process(ref_wb, new_wb, pdf_bytes: bytes) -> bytes:
    ref_map = build_ref_map(ref_wb)
    full_text = extract_pdf_text(pdf_bytes)
    ws = new_wb.active

    pool_passage = find_passage(full_text, POOL_KW, window=1000, start_at_kw=True)

    for row in ws.iter_rows(min_row=1):
        if len(row) < 8:
            continue

        desc = row[1].value
        col_c = row[2].value
        e_cell = row[4]
        f_cell = row[5]
        g_cell = row[6]
        h_cell = row[7]

        # Skip non-data rows
        if not desc or not isinstance(desc, str):
            continue
        if desc in SKIP_ROWS:
            continue
        if desc.startswith("'") or desc.startswith("Pg ") or desc.startswith("Page "):
            continue
        # Skip if already filled
        if e_cell.value is not None:
            continue

        row_num = row[1].row

        # ── Formula rows: find percentage in PDF by value, verify by context ──
        if desc in FORMULA_ROWS and isinstance(col_c, (int, float)) and col_c != 0:
            kws = SEARCH_KW.get(desc, [desc])
            passage, pct_val, prec = find_pct_passage(full_text, col_c, kws)

            e_cell.value = passage or ref_map.get(desc, {}).get("lang", "")
            if pct_val is not None:
                f_cell.value = pct_val
                h_cell.value = prec
                g_cell.value = f"=ROUND(C{row_num},H{row_num})%-F{row_num}"
            else:
                f_cell.value = "N/A"
                g_cell.value = "Handtie"
            continue

        # ── Handtie rows ─────────────────────────────────────────────────────
        if desc in FIXED_LANGUAGE:
            e_cell.value = FIXED_LANGUAGE[desc]

        elif desc in POOL_DESC_ROWS:
            e_cell.value = pool_passage or ref_map.get(desc, {}).get("lang", "")

        elif desc in SEARCH_KW:
            kws = SEARCH_KW[desc]
            # Start passage at keyword for cleaner sentence start
            e_cell.value = find_passage(full_text, kws, start_at_kw=True) or ref_map.get(desc, {}).get("lang", "")

        else:
            ref = ref_map.get(desc)
            if ref:
                lang = ref["lang"]
                # Use first ~5 words of reference language as search seed
                seed = " ".join(lang.split()[:5])
                found = find_passage(full_text, [seed]) if seed else None
                e_cell.value = found or lang
            else:
                e_cell.value = find_passage(full_text, [desc]) or ""

        f_cell.value = "N/A"
        g_cell.value = "Handtie"

    buf = io.BytesIO()
    new_wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ──────────────────────────────────────────────────────────────

st.set_page_config(page_title="PPM Ack Letter Bot", page_icon="📋", layout="centered")

st.title("📋 PPM Ack Letter Support Bot")
st.caption(
    "Upload the three files below. The bot fills PPM Language (col E), "
    "PPM Number (col F), Etie formula (col G), and Precision (col H)."
)

ref_file = st.file_uploader("1. Reference deal Excel (previous deal)", type=["xlsx", "xls"])
new_file = st.file_uploader("2. New deal Excel (current deal support file)", type=["xlsx", "xls"])
pdf_file = st.file_uploader("3. New deal PDF (Private Placement Memorandum)", type=["pdf"])

if ref_file and new_file and pdf_file:
    with st.spinner("Reading files and extracting from PDF…"):
        ref_wb = openpyxl.load_workbook(io.BytesIO(ref_file.read()), data_only=True)
        new_wb = openpyxl.load_workbook(io.BytesIO(new_file.read()))
        result = process(ref_wb, new_wb, pdf_file.read())

    name = new_file.name.rsplit(".", 1)[0]
    st.success("Done! Download your filled Excel below.")
    st.download_button(
        label="⬇️  Download Filled Excel",
        data=result,
        file_name=f"{name}_filled.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
