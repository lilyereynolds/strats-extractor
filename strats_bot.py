"""
PPM Strats PDF → Excel Bot
Upload any CHASE-style PPM strats PDF and download all tables as Excel.
Run: streamlit run strats_bot.py
"""

import io
import re
import pdfplumber
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import streamlit as st

# ── Column x-boundaries (divide each PDF page into 10 columns) ───────────────
COL_BOUNDS = [295, 345, 430, 485, 535, 610, 660, 710, 755]

# Full multiline column headers matching the example Excel
FULL_COL_HEADERS = [
    "Number of\nLoans",
    "Aggregate Stated\nPrincipal Balance\n($)",
    "Aggregate\nStated\nPrincipal\nBalance (%)",
    "Weighted\nAverage\nCurrent\nMortgage\nRate (%)",
    "Average\nStated\nPrincipal\nBalance ($)",
    "Weighted\nAverage\nOriginal\nCredit\nScore",
    "Weighted\nAverage\nCurrent\nCredit\nScore",
    "Weighted\nAverage\nOriginal\nLTV (%)",
    "Weighted\nAverage\nOriginal\nCombined\nLTV (%)",
]

HEADER_WORDS = {
    "weighted", "aggregate", "average", "stated", "current",
    "number", "principal", "balance", "mortgage", "credit",
    "original", "combined", "loans", "rate", "score",
}

# Row types used for formatting
TITLE    = "TITLE"
SECTION  = "SECTION"
HEADER   = "HEADER"
DATA     = "DATA"
TOTAL    = "TOTAL"
FOOTNOTE = "FOOTNOTE"
FOOTER   = "FOOTER"
EMPTY    = "EMPTY"


# ── Parsing helpers ───────────────────────────────────────────────────────────

def assign_col(x: float) -> int:
    for i, b in enumerate(COL_BOUNDS):
        if x < b:
            return i
    return 9


def parse_row(row_words: list) -> dict:
    cols: dict[int, str] = {}
    for w in row_words:
        c = assign_col(w["x0"])
        t = w["text"].strip()
        if t:
            cols[c] = (cols[c] + " " + t).strip() if c in cols else t
    return cols


def is_header_frag(cols: dict) -> bool:
    """True for the 4 partial header rows (no col-0, ≥3 cols, all header keywords)."""
    if 0 in cols or len(cols) < 3:
        return False
    return all(any(kw in v.lower() for kw in HEADER_WORDS) for v in cols.values())


def is_header_label(cols: dict) -> bool:
    """True for the last header row that has the col-0 label + 'Loans'/($) etc."""
    if 0 not in cols:
        return False
    c1 = cols.get(1, "").lower()
    c2 = cols.get(2, "").lower()
    return "loans" in c1 or "($)" in c2 or "($)" in c1


def is_data_row(cols: dict) -> bool:
    if 0 not in cols:
        return False
    raw = cols.get(1, "").replace(",", "").replace(".", "").strip()
    return raw.isdigit() and len(raw) > 0


def is_section_label(text: str) -> bool:
    """Bold section titles: short, starts uppercase, not a footnote/continuation."""
    t = text.strip()
    if not t or re.match(r"^\(\d+\)", t):
        return False
    if t[0].islower():
        return False
    if len(t) > 80 and t.endswith("."):   # long sentence = footnote fragment
        return False
    return True


def to_num(s: str):
    s = str(s).strip().replace(",", "")
    try:
        return int(s) if "." not in s else float(s)
    except ValueError:
        return s


def extract_footnote_number(text: str):
    """Pull the first meaningful number from a footnote, skipping the leading (N) reference."""
    clean = re.sub(r'^\(\d+\)\s*', '', text.strip())
    matches = re.findall(r'\$?([\d,]+(?:\.\d+)?)%?', clean)
    for m in matches:
        v = m.replace(",", "")
        try:
            return int(v) if "." not in v else float(v)
        except ValueError:
            continue
    return None


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_pdf(pdf_bytes: bytes) -> list[tuple[str, list]]:
    """Return list of (row_type, row_data) for every meaningful row in the PDF."""
    output: list[tuple[str, list]] = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pg_idx, page in enumerate(pdf.pages):
            words = page.extract_words(keep_blank_chars=True)

            y_groups: dict[int, list] = {}
            for w in words:
                yk = round(w["top"] / 3) * 3
                y_groups.setdefault(yk, []).append(w)

            in_header = False

            for y in sorted(y_groups):
                row_words = sorted(y_groups[y], key=lambda w: w["x0"])
                all_text = " ".join(
                    w["text"].strip() for w in row_words if w["text"].strip()
                )
                if not all_text.strip():
                    continue

                cols = parse_row(row_words)
                if not cols:
                    continue

                # ── Page footer (A-N marker) ─────────────────────────────────
                if re.match(r"^A-\d+\s*$", all_text.strip()):
                    output.append((FOOTER, [all_text.strip()] + [None] * 9))
                    output.append((EMPTY,  [None] * 10))
                    in_header = False
                    continue

                # ── Header fragments — skip (reconstructed below) ────────────
                if is_header_frag(cols):
                    in_header = True
                    continue

                # ── Header label row (last row of 5-row header block) ────────
                if in_header and is_header_label(cols):
                    in_header = False
                    col1 = cols.get(0, "").strip()
                    output.append((HEADER, [col1] + FULL_COL_HEADERS))
                    continue

                # ── Centered page title (page 1 only) ────────────────────────
                if pg_idx == 0 and 0 not in cols and not is_header_frag(cols):
                    title = " ".join(cols.values()).strip()
                    if title:
                        output.append((TITLE, [title] + [None] * 9))
                    continue

                # ── Col-0-only rows: section labels OR footnote text ─────────
                if set(cols.keys()) == {0}:
                    text = cols[0].strip()
                    if not text:
                        continue
                    if is_section_label(text):
                        output.append((SECTION, [text] + [None] * 9))
                    else:
                        num = extract_footnote_number(text)
                        row = [text, num] + [None] * 8
                        output.append((FOOTNOTE, row))
                    continue

                # ── Data rows ────────────────────────────────────────────────
                if is_data_row(cols):
                    row = [None] * 10
                    row[0] = cols.get(0, "").strip()
                    for ci in range(1, 10):
                        v = cols.get(ci, "").strip()
                        if v:
                            row[ci] = to_num(v)
                    rtype = TOTAL if row[0] == "Total:" else DATA
                    output.append((rtype, row))
                    continue

                # ── Header label without preceding frags ─────────────────────
                if is_header_label(cols):
                    col1 = cols.get(0, "").strip()
                    output.append((HEADER, [col1] + FULL_COL_HEADERS))

    return output


# ── Post-processor ───────────────────────────────────────────────────────────

def post_process_rows(tagged_rows: list[tuple[str, list]]) -> list[tuple[str, list]]:
    # Step 1: merge footnote continuation lines into one row per footnote
    merged = []
    i = 0
    while i < len(tagged_rows):
        rtype, row_data = tagged_rows[i]
        if rtype == FOOTNOTE:
            text = row_data[0]
            num = row_data[1]
            j = i + 1
            while j < len(tagged_rows) and tagged_rows[j][0] == FOOTNOTE and not re.match(r'^\(\d+\)', tagged_rows[j][1][0]):
                text = text + " " + tagged_rows[j][1][0]
                j += 1
            merged.append((FOOTNOTE, [text, num] + [None] * 8))
            i = j
        else:
            merged.append((rtype, row_data))
            i += 1

    # Step 2: insert one blank row after the last row of each table
    # (after TOTAL if no footnotes follow, or after the last FOOTNOTE)
    result = []
    for i, (rtype, row_data) in enumerate(merged):
        result.append((rtype, row_data))
        if rtype in (TOTAL, FOOTNOTE):
            next_type = None
            has_footer = False
            for j in range(i + 1, len(merged)):
                t = merged[j][0]
                if t == FOOTER:
                    has_footer = True
                if t not in (EMPTY, FOOTER):
                    next_type = t
                    break
            if not has_footer and next_type in (HEADER, SECTION):
                result.append((EMPTY, [None] * 10))
    return result


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(tagged_rows: list[tuple[str, list]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Fonts matching example: Times New Roman 8pt, bold via separate font name
    TNR_BOLD  = Font(name="Times New Roman", size=8, bold=True)
    TNR_REG   = Font(name="Times New Roman", size=8)
    TNR_9     = Font(name="Times New Roman", size=9)
    TNR_TITLE = Font(name="Times New Roman", size=11, bold=True)

    WRAP_TOP  = Alignment(wrap_text=True, vertical="top")
    CENTER    = Alignment(horizontal="center", vertical="center")

    for rtype, row_data in tagged_rows:
        ws.append(row_data)
        r = ws.max_row

        if rtype == TITLE:
            # Merge A:J, center, bold title
            ws.merge_cells(f"A{r}:J{r}")
            ws[f"A{r}"].value = row_data[0]
            ws[f"A{r}"].font = TNR_TITLE
            ws[f"A{r}"].alignment = CENTER

        elif rtype == SECTION:
            ws.cell(r, 1).font = TNR_BOLD

        elif rtype == HEADER:
            for c in range(1, 11):
                cell = ws.cell(r, c)
                cell.font = TNR_BOLD
                cell.alignment = WRAP_TOP
            ws.row_dimensions[r].height = 53.5

        elif rtype == TOTAL:
            for c in range(1, 11):
                ws.cell(r, c).font = TNR_BOLD

        elif rtype == DATA:
            for c in range(1, 11):
                ws.cell(r, c).font = TNR_REG

        elif rtype == FOOTNOTE:
            ws.cell(r, 1).font = TNR_REG

        elif rtype == FOOTER:
            ws.cell(r, 1).font = TNR_9

        # EMPTY rows: no font override needed

    # Column widths matching example Excel
    col_widths = {
        "A": 30.63, "B":  7.09, "C": 28.63, "D":  8.27, "E":  6.73,
        "F": 10.09, "G":  6.45, "H": 13.0,  "I": 13.0,  "J":  6.73,
    }
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="PDF Strats Extractor",
    page_icon="📊",
    layout="centered",
)

st.title("📊 PDF Strats → Excel")
st.caption(
    "Upload a PDF Strats file. "
    "All tables and footnotes are extracted into a single Excel workbook."
)

uploaded = st.file_uploader(
    "Drop your PDF strats file here",
    type="pdf",
    label_visibility="collapsed",
)

if uploaded:
    with st.spinner(f"Extracting from **{uploaded.name}**…"):
        tagged = post_process_rows(parse_pdf(uploaded.read()))

    rows_only = [r for _, r in tagged]
    data_rows = [
        r for t, r in tagged
        if t == DATA and any(isinstance(v, (int, float)) for v in r[1:])
    ]
    n_headers = sum(1 for t, _ in tagged if t == HEADER)

    if not data_rows:
        st.error("No data tables found. Make sure this is a PPM strats PDF.")
    else:
        st.success(
            f"Extracted **{len(data_rows)}** data rows across "
            f"**{n_headers}** strat tables."
        )

        xlsx = build_excel(tagged)
        stem = uploaded.name.rsplit(".", 1)[0].replace(" ", "_")

        st.download_button(
            label="⬇️  Download Excel",
            data=xlsx,
            file_name=f"{stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
